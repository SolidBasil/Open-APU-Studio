"""
exportar_excel.py
=================
Exportación a Excel (.xlsx) de la cuantificación automática por capa.

Genera un workbook con hoja "Quantities by layer" y opcionalmente
"Count by block". Defensa contra inyección de fórmulas en celdas.
"""

from __future__ import annotations

import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .auto_quantify import LayerQuantity

_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FONT = Font(bold=True)


def _neutralise_formula(value: str) -> str:
    """Defensa OWASP contra inyección de fórmulas en celdas."""
    if value and re.match(r"^[=+\-@\t\r\n]", value):
        return f"'{value}"
    return value


def export_quantify_to_excel(
    layer_quantities: list[LayerQuantity],
    *,
    drawing_name: str = "drawing",
    count_total: int | None = None,
    by_block: list[dict] | None = None,
    output_path: str | None = None,
) -> str:
    """Exporta cuantificación por capa a .xlsx.

    Returns:
        Ruta del archivo generado.
    """
    wb = Workbook()

    # ── Hoja: Quantities by layer ─────────────────────────────
    ws = wb.active
    ws.title = "Quantities by layer"
    headers = ["Layer", "Measure", "Quantity", "Unit", "Area (m²)", "Length (m)", "Entities"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 10

    total_area = 0.0
    total_length = 0.0
    total_entities = 0
    r3 = lambda n: round(n * 1000) / 1000

    for row in layer_quantities:
        total_area += row.area
        total_length += row.length
        total_entities += row.count
        ws.append([
            _neutralise_formula(row.layer),
            row.primary,
            r3(row.quantity),
            row.unit,
            r3(row.area) if row.area > 0 else "",
            r3(row.length) if row.length > 0 else "",
            row.count,
        ])

    # Total row
    ws.append(["TOTAL", "", "", "", r3(total_area), r3(total_length), total_entities])
    total_row = ws[ws.max_row]
    for cell in total_row:
        cell.font = _TOTAL_FONT

    if count_total and count_total > 0:
        ws.append(["Manual count items", "", count_total, "nr", "", "", ""])
        ws[ws.max_row][0].font = Font(italic=True)

    # ── Hoja: Count by block ──────────────────────────────────
    if by_block:
        ws_b = wb.create_sheet("Count by block")
        ws_b.append(["Block", "Count"])
        for cell in ws_b[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
        ws_b.column_dimensions["A"].width = 30
        ws_b.column_dimensions["B"].width = 10

        bt = 0
        for b in by_block:
            bt += b.get("count", 0)
            ws_b.append([_neutralise_formula(b.get("name", "")), b.get("count", 0)])
        ws_b.append(["TOTAL", bt])
        ws_b[ws_b.max_row][0].font = _TOTAL_FONT

    slug = re.sub(r"[^\w]+", "_", drawing_name).strip("_").lower() or "drawing"
    out = output_path or f"dwg-quantities-{slug}.xlsx"
    wb.save(out)
    return os.path.abspath(out)
