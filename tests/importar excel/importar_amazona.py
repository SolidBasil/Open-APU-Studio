"""
Importa CATALOGO_DE_CONCEPTOS_AMPLIACIÓN_AMAZONA_VINCULADO_1.xlsx
a un proyecto .db de Open APU Studio (presupuesto + generadores, sin matrices).
"""
import sys
from pathlib import Path
import sqlite3
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    print("openpyxl no instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).parent))
from backend.database.core import generar_hash

EXCEL = Path(__file__).parent / "CATALOGO_DE_CONCEPTOS_AMPLIACIÓN_AMAZONA_VINCULADO_1.xlsx"
SCHEMA = Path(__file__).parent / "backend" / "database" / "schema.sql"

TIPO_ID_CONCEPTO_COMPUESTO = 32  # tipos_insumo.clave = 'concepto'

if not EXCEL.exists():
    print(f"No se encuentra: {EXCEL}")
    sys.exit(1)
if not SCHEMA.exists():
    print(f"No se encuentra: {SCHEMA}")
    sys.exit(1)


def aplicar_schema(conn):
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    sql = SCHEMA.read_text(encoding="utf-8")
    conn.executescript(sql)
    conn.execute("UPDATE estructura_presupuesto SET es_extra = 0 WHERE es_extra IS NULL")
    conn.execute("""UPDATE proyectos SET
        horas_dia=8, capturar_rendimientos=0, unidad_cantidad_agrup=0,
        tipo_cambio=1.0, reporte_version='1.0'""")
    conn.commit()


def calcular_nivel(tipo_col_a, ultimo_nivel):
    """Nivel jerárquico leído DIRECTO de la columna A del Excel
    (Capítulo / Subcapítulo / Concepto) — nunca se infiere de la clave
    (columna C, 'clave opus', que no se importa) ni de ningún otro campo.

    Si la celda viene vacía, se asume continuación del último nivel visto
    (en este archivo eso solo ocurre entre renglones de Concepto).
    """
    t = str(tipo_col_a or "").strip().capitalize()
    if t == "Capítulo":
        return 0
    if t == "Subcapítulo":
        return 1
    if t == "Concepto":
        return 2
    return ultimo_nivel if ultimo_nivel is not None else 2


def obtener_o_crear_insumo_compuesto(conn, proyecto_id, descripcion, unidad):
    """Crea (o reutiliza si ya existe por hash) un insumo 'Concepto compuesto'
    para representar un concepto del presupuesto, en vez de escribir la
    descripción directo en estructura_presupuesto.descripcion. Así el árbol
    resuelve descripcion/unidad/precio via JOIN a insumos, igual que cualquier
    concepto creado a mano en la app."""
    hash_ = generar_hash(descripcion)
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM insumos WHERE proyecto_id = ? AND hash = ?",
        (proyecto_id, hash_),
    )
    row = cur.fetchone()
    if row:
        print(f"    [insumo reutilizado] id={row['id']}  {descripcion[:60]!r}")
        return row["id"]

    cur.execute(
        """INSERT INTO insumos
           (proyecto_id, hash, tipo_id, es_compuesto, descripcion, unidad,
            costo_mn, costo_me, costo_directo, costo_final)
           VALUES (?,?,?,1,?,?,0,0,0,0)""",
        (proyecto_id, hash_, TIPO_ID_CONCEPTO_COMPUESTO, descripcion, unidad),
    )
    nuevo_id = cur.lastrowid
    print(f"    [insumo creado]     id={nuevo_id}  {descripcion[:60]!r}")
    return nuevo_id


def importar_catalogo(conn, proyecto_id):
    """Lee hoja 'Catalogo de conceptos' y construye el árbol.

    El nivel viene de la columna A (Capítulo/Subcapítulo/Concepto). La
    columna C (clave OPUS) se ignora por completo: no se guarda como wbs
    ni se usa para nada estructural. El wbs que ve la app se genera aquí
    mismo, de forma puramente jerárquica (1, 1.1, 1.1.1, ...).
    """
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb["Catalogo de conceptos "]

    capitulo_actual = None      # nodo (dict) del capítulo abierto
    subcapitulo_actual = None   # nodo (dict) del subcapítulo abierto
    ultimo_nivel = None
    contador_cap = 0
    contador_sub = 0
    contador_con = 0
    orden = 0
    nodos = []       # (parent_node_o_None, nivel, wbs_generado, desc, unidad, cantidad, orden)
    siguiente_key = 0

    def nueva_key():
        nonlocal siguiente_key
        siguiente_key += 1
        return siguiente_key

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        desc = str(row[3] or "").strip() if len(row) > 3 else ""
        if not desc:
            continue

        nivel = calcular_nivel(row[0] if len(row) > 0 else None, ultimo_nivel)
        ultimo_nivel = nivel

        unidad = str(row[4] or "").strip() if len(row) > 4 else ""
        cantidad_raw = row[5] if len(row) > 5 else None
        cantidad = 0.0
        if cantidad_raw is not None:
            try:
                cantidad = float(cantidad_raw)
            except (ValueError, TypeError):
                pass

        if nivel == 0:
            contador_cap += 1
            contador_sub = 0
            wbs = str(contador_cap)
            nodo = {"key": nueva_key(), "wbs": wbs}
            capitulo_actual = nodo
            subcapitulo_actual = None
            orden += 1
            nodos.append((None, nivel, wbs, desc, "", 0.0, orden, nodo["key"]))
        elif nivel == 1:
            contador_sub += 1
            base = capitulo_actual["wbs"] if capitulo_actual else str(contador_cap or 1)
            wbs = f"{base}.{contador_sub}"
            nodo = {"key": nueva_key(), "wbs": wbs}
            subcapitulo_actual = nodo
            orden += 1
            nodos.append((capitulo_actual, nivel, wbs, desc, "", 0.0, orden, nodo["key"]))
        else:
            contador_con += 1
            padre = subcapitulo_actual or capitulo_actual
            base = padre["wbs"] if padre else "0"
            wbs = f"{base}.{contador_con}"
            orden += 1
            nodos.append((padre, nivel, wbs, desc, unidad, cantidad, orden, None))

    wb.close()

    # Insertar nodos obteniendo ids, indexados por la 'key' interna (no por
    # nada del Excel) para resolver padre_id.
    id_por_key = {}
    cur = conn.cursor()
    n_insumos_creados = 0
    for parent, nivel, wbs, desc, unidad, cantidad, ord_, key in nodos:
        padre_id = id_por_key.get(parent["key"]) if parent else None

        if nivel == 2:
            insumo_id = obtener_o_crear_insumo_compuesto(conn, proyecto_id, desc, unidad)
            n_insumos_creados += 1
            cur.execute(
                """INSERT INTO estructura_presupuesto
                   (proyecto_id, padre_id, nivel, wbs, descripcion,
                    cantidad, total, orden, tipo, insumo_id)
                   VALUES (?,?,?,?,'',?,0,?,?,?)""",
                (proyecto_id, padre_id, nivel, wbs,
                 cantidad, ord_, "concepto", insumo_id),
            )
            new_id = cur.lastrowid
            print(f"    concepto id={new_id} wbs={wbs}  → insumo_id={insumo_id}")
        else:
            cur.execute(
                """INSERT INTO estructura_presupuesto
                   (proyecto_id, padre_id, nivel, wbs, descripcion,
                    cantidad, total, orden, tipo)
                   VALUES (?,?,?,?,?,?,0,?,?)""",
                (proyecto_id, padre_id, nivel, wbs, desc,
                 cantidad, ord_, "capitulo"),
            )
            new_id = cur.lastrowid

        if key is not None:
            id_por_key[key] = new_id

    conn.commit()
    print(f"  {len(nodos)} nodos insertados ({n_insumos_creados} conceptos → {n_insumos_creados} insumos esperados)")

    cur.execute("SELECT COUNT(*) FROM insumos WHERE proyecto_id = ?", (proyecto_id,))
    n_insumos_reales = cur.fetchone()[0]
    cur.execute(
        "SELECT COUNT(*) FROM estructura_presupuesto "
        "WHERE proyecto_id = ? AND tipo = 'concepto' AND insumo_id IS NULL",
        (proyecto_id,),
    )
    n_sin_ligar = cur.fetchone()[0]
    print(f"  insumos en DB: {n_insumos_reales}  |  conceptos sin insumo_id: {n_sin_ligar}")
    if n_sin_ligar:
        print("  ¡ATENCIÓN! Hay conceptos sin insumo ligado, revisar arriba.")

    # Recalcular totales
    recalcular_totales(conn, proyecto_id)

    # id_map ya no se indexa por clave (columna C ignorada); se devuelve
    # vacío porque importar_generadores empareja por texto de descripción,
    # no por este diccionario (ver ese código más abajo).
    return {}


def recalcular_totales(conn, proyecto_id):
    cur = conn.cursor()
    cur.execute(
        "SELECT id, nivel, cantidad, padre_id FROM estructura_presupuesto "
        "WHERE proyecto_id = ? AND activo = 1 ORDER BY id",
        (proyecto_id,),
    )
    nodos = cur.fetchall()
    totals = {}
    for row in nodos:
        nid = row["id"]
        nivel = row["nivel"]
        cant = float(row["cantidad"] or 0)
        padre_id = row["padre_id"]
        if nivel == 2:
            totals[nid] = cant
        else:
            totals[nid] = 0

    updated = {}
    for row in reversed(nodos):
        nid = row["id"]
        pid = row["padre_id"]
        total = totals.get(nid, 0)
        if pid and pid in totals:
            totals[pid] = totals.get(pid, 0) + total

    for nid, total in totals.items():
        cur.execute("UPDATE estructura_presupuesto SET total = ? WHERE id = ?", (total, nid))
    conn.commit()


def importar_generadores(conn, proyecto_id, id_map):
    """Lee hoja 'Generadores' y crea generadores + renglones."""
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb["Generadores"]

    cur = conn.cursor()
    gen_count = 0
    renglon_count = 0

    rows = list(ws.iter_rows(values_only=True))
    i = 0
    while i < len(rows):
        row = rows[i]
        # Buscar bloque: encontrar CONCEPTO:
        if row and len(row) > 2 and str(row[0] or "").strip() == "CONCEPTO:":
            desc = str(row[1] or "").strip() if len(row) > 1 else ""
            total_raw = row[8] if len(row) > 8 else None
            total = float(total_raw) if total_raw and isinstance(total_raw, (int, float)) else 0.0

            # Encontrar concepto en id_map por descripción
            concepto_id = None
            for clave, nid in id_map.items():
                if clave and clave.replace("-", "").strip():
                    # Buscar coincidencia aproximada de texto
                    pass
            # Buscar por coincidencia de texto en la descripción del insumo
            # ligado (estructura_presupuesto.descripcion queda vacía para
            # conceptos desde que se creó el insumo compuesto — ver arriba).
            cur.execute(
                "SELECT ep.id FROM estructura_presupuesto ep "
                "JOIN insumos i ON i.id = ep.insumo_id "
                "WHERE ep.proyecto_id = ? AND ep.tipo = 'concepto' "
                "AND i.descripcion LIKE ? LIMIT 1",
                (proyecto_id, f"%{desc[:40]}%"),
            )
            match = cur.fetchone()
            if match:
                concepto_id = match["id"]

            # Crear generador
            cur.execute(
                "INSERT INTO generadores (proyecto_id, concepto_id, nombre, cantidad_total) "
                "VALUES (?,?,?,?)",
                (proyecto_id, concepto_id, desc[:100], 0.0),
            )
            generador_id = cur.lastrowid
            gen_count += 1

            # Saltar a la fila de encabezados de columnas
            i += 1
            while i < len(rows):
                r = rows[i]
                if r and len(r) > 2 and str(r[0] or "").strip() == "CONCEPTO:":
                    break
                if r and len(r) > 3 and str(r[3] or "").strip().upper() in ("LARGO",):
                    i += 1
                    break
                i += 1

            # Leer renglones hasta que haya una fila vacía o nuevo CONCEPTO
            while i < len(rows):
                r = rows[i]
                if not r or all(v is None for v in r):
                    i += 1
                    break
                if r[0] and str(r[0]).strip() == "CONCEPTO:":
                    break

                eje = str(r[0] or "").strip()
                tramo = str(r[1] or "").strip() if len(r) > 1 else ""
                unidad = str(r[2] or "").strip() if len(r) > 2 else ""

                def to_float(v):
                    if v is None:
                        return None
                    try:
                        return float(v)
                    except (ValueError, TypeError):
                        return None

                largo = to_float(r[3] if len(r) > 3 else None)
                ancho = to_float(r[4] if len(r) > 4 else None)
                alto = to_float(r[5] if len(r) > 5 else None)
                veces = to_float(r[6] if len(r) > 6 else None)

                subtotal = 0.0
                if len(r) > 7 and r[7] is not None:
                    try:
                        subtotal = float(r[7])
                    except (ValueError, TypeError):
                        pass

                cur.execute(
                    """INSERT INTO generador_renglones
                       (generador_id, orden, eje, tramo, veces, largo, ancho, alto, subtotal)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (generador_id, renglon_count + 1, eje, tramo, veces or 1.0,
                     largo, ancho, alto, subtotal),
                )
                renglon_count += 1
                i += 1
            continue
        i += 1

    conn.commit()

    # Recalcular cantidades de cada generador y su concepto
    cur.execute("SELECT id, concepto_id FROM generadores WHERE proyecto_id = ? AND activo = 1",
                (proyecto_id,))
    for g in cur.fetchall():
        gid = g["id"]
        cid = g["concepto_id"]
        cur.execute(
            "SELECT COALESCE(SUM(subtotal), 0) FROM generador_renglones "
            "WHERE generador_id = ? AND activo = 1",
            (gid,),
        )
        ct = float(cur.fetchone()[0])
        cur.execute("UPDATE generadores SET cantidad_total = ? WHERE id = ?", (ct, gid))
        if cid:
            cur.execute(
                "SELECT COALESCE(SUM(cantidad_total), 0) FROM generadores "
                "WHERE concepto_id = ? AND activo = 1",
                (cid,),
            )
            total_cant = float(cur.fetchone()[0])
            cur.execute(
                "UPDATE estructura_presupuesto SET cantidad = ? WHERE id = ?",
                (total_cant, cid),
            )

    conn.commit()

    # Recalcular totales finales
    recalcular_totales(conn, proyecto_id)
    
    print(f"  {gen_count} generadores, {renglon_count} renglones")


def main():
    nombre = "AMAZONA_CATALOGO"
    ruta_db = Path(__file__).parent / "backend" / "datos_usuario" / "proyectos" / f"{nombre}.db"

    if ruta_db.exists():
        ruta_db.unlink()

    print(f"Creando proyecto: {ruta_db}")

    conn = sqlite3.connect(str(ruta_db))
    conn.row_factory = sqlite3.Row

    aplicar_schema(conn)
    print("Schema aplicado.")

    # Crear proyecto
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO proyectos (nombre, obra_descripcion) VALUES (?,?)",
        (nombre, "AMPLIACIÓN AMAZONA"),
    )
    proyecto_id = cur.lastrowid

    # Insertar usuario por defecto
    cur.execute(
        "INSERT OR IGNORE INTO usuarios (id, nombre, email) VALUES (1,'admin','admin@local')",
    )
    conn.commit()
    print(f"Proyecto id={proyecto_id} creado.")

    # Importar catálogo
    print("Importando catálogo de conceptos...")
    id_map = importar_catalogo(conn, proyecto_id)
    print(f"  {len(id_map)} claves mapeadas.")

    # Importar generadores
    print("Importando generadores...")
    importar_generadores(conn, proyecto_id, id_map)

    conn.close()
    print(f"\nProyecto creado: {ruta_db}")


if __name__ == "__main__":
    main()