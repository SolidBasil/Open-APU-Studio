from __future__ import annotations

import re
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_TITLE_FONT = Font(bold=True, size=11)
_NORMAL_FONT = Font(size=10)
_TOTAL_FONT = Font(bold=True, size=10)
_CONCEPTO_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
_LABEL_FONT = Font(bold=True, size=10)
_HOJA_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
_THICK_TOP = Border(top=Side(style="medium", color="4472C4"))
_THICK_BOTTOM = Border(bottom=Side(style="medium", color="4472C4"))
_THICK_LEFT = Border(left=Side(style="medium", color="4472C4"))
_THICK_RIGHT = Border(right=Side(style="medium", color="4472C4"))

_NUM_FMT = '#,##0.00'
_NUM_FMT_4 = '#,##0.0000'


def _neutralise(value: str | None) -> str | None:
    if value and isinstance(value, str) and re.match(r"^[=+\-@\t\r\n]", value):
        return f"'{value}"
    return value


def _wbs_sort_key(wbs: str | None) -> tuple:
    """Sort key para WBS tipo '1.2.10' → (1, 2, 10) orden numérico."""
    if not wbs:
        return (0,)
    parts = wbs.replace("-", ".").split(".")
    result = []
    for p in parts:
        try:
            result.append(int(p))
        except ValueError:
            result.append(0)
    return tuple(result)


def _style_header(ws, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _auto_width(ws, min_width: int = 8, max_width: int = 50):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            val = str(cell.value or "")
            lengths.append(min(len(val), max_width))
        best = max(lengths) + 2 if lengths else min_width
        ws.column_dimensions[col_letter].width = max(min_width, min(best, max_width))


def _write_catalogo_sheet(ws, conn: sqlite3.Connection, proyecto_id: int):
    headers = ["Tipo", "CLAVE", "DESCRIPCIÓN", "UNIDAD", "CANTIDAD",
               "PRECIO UNITARIO", "IMPORTE"]
    ws.append(headers)
    _style_header(ws, len(headers))

    cur = conn.cursor()
    cur.execute("""
        SELECT
            n.id, n.padre_id, n.wbs, n.nivel, n.tipo,
            n.cantidad, n.total,
            CASE WHEN n.tipo = 'concepto'
                THEN COALESCE(i.descripcion, n.descripcion)
                ELSE n.descripcion
            END AS descripcion,
            i.unidad,
            i.costo_final AS precio_unitario
        FROM estructura_presupuesto n
        LEFT JOIN insumos i ON i.id = n.insumo_id
        WHERE n.proyecto_id = ? AND n.activo = 1 AND n.es_extra = 0
    """, (proyecto_id,))

    def _tipo_label(row):
        if row["tipo"] == "concepto":
            return "Concepto"
        nivel = row["nivel"] or 0
        if nivel == 0:
            return "Capítulo"
        if nivel == 1:
            return "Subcapítulo"
        return f"Capítulo (n.{nivel})"

    rows = sorted(cur.fetchall(), key=lambda r: _wbs_sort_key(r["wbs"]))
    for row in rows:
        tipo = _tipo_label(row)
        ws.append([
            tipo,
            _neutralise(row["wbs"]),
            _neutralise(row["descripcion"]),
            row["unidad"] or "",
            round(row["cantidad"], 4) if row["cantidad"] else 0.0,
            round(row["precio_unitario"], 2) if row["precio_unitario"] else 0.0,
            round(row["total"], 2) if row["total"] else 0.0,
        ])

    CAP_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    CAP_FONT = Font(bold=True, color="FFFFFF", size=10)
    SUB_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    SUB_FONT = Font(bold=True, size=10)

    for row_idx in range(2, ws.max_row + 1):
        tipo_val = str(ws.cell(row=row_idx, column=1).value or "")
        is_capitulo = tipo_val == "Capítulo"
        is_sub = tipo_val == "Subcapítulo"
        for col_idx in [5, 6, 7]:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = _NUM_FMT
            cell.border = _THIN_BORDER
        cell = ws.cell(row=row_idx, column=5)
        cell.number_format = _NUM_FMT_4
        if is_capitulo or is_sub:
            fill = CAP_FILL if is_capitulo else SUB_FILL
            font = CAP_FONT if is_capitulo else SUB_FONT
            for col_idx in range(1, 8):
                c = ws.cell(row=row_idx, column=col_idx)
                c.fill = fill
                c.font = font
                c.border = _THIN_BORDER
    _auto_width(ws)


def _write_generadores_sheet(ws, conn: sqlite3.Connection, proyecto_id: int):
    cur = conn.cursor()
    from datetime import datetime

    # ── Datos del proyecto ──
    cur.execute("""
        SELECT nombre, obra_descripcion, obra_domicilio, obra_ciudad, obra_estado
        FROM proyectos WHERE id = ?
    """, (proyecto_id,))
    proy = cur.fetchone()
    obra_nombre = (proy["nombre"] or "") if proy else ""
    obra_ubicacion = ""
    if proy:
        partes = [proy["obra_domicilio"], proy["obra_ciudad"], proy["obra_estado"]]
        obra_ubicacion = ", ".join(p for p in partes if p)
    fecha_str = datetime.now().strftime("%d/%m/%Y")

    NC = 8  # columnas sin Unidad
    row_num = 1

    # ── Encabezado tipo hoja de generadores ──
    for r in range(row_num, row_num + 2):
        for c in range(1, NC + 1):
            ws.cell(row=r, column=c).fill = _HOJA_FILL
    ws.cell(row=row_num, column=1, value="Obra:").font = _LABEL_FONT
    ws.cell(row=row_num, column=2, value=obra_nombre)
    ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
    ws.cell(row=row_num, column=7, value="Fecha:").font = _LABEL_FONT
    ws.cell(row=row_num, column=8, value=fecha_str)
    row_num += 1
    ws.cell(row=row_num, column=1, value="Ubicación:").font = _LABEL_FONT
    ws.cell(row=row_num, column=2, value=obra_ubicacion or "")
    ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
    ws.cell(row=row_num, column=7, value="Hoja No:").font = _LABEL_FONT
    ws.cell(row=row_num, column=8, value=1)
    row_num += 2  # espacio antes del primer concepto

    # ── Generadores ──
    cur.execute("""
        SELECT
            g.id, g.nombre, g.cantidad_total, g.concepto_id, g.notas,
            g.unidad AS gen_unidad,
            CASE WHEN ep.tipo = 'concepto'
                THEN COALESCE(i.descripcion, ep.descripcion)
                ELSE ep.descripcion
            END AS concepto_descripcion,
            COALESCE(i.unidad, g.unidad) AS unidad,
            ep.wbs
        FROM generadores g
        JOIN estructura_presupuesto ep ON ep.id = g.concepto_id
        LEFT JOIN insumos i ON i.id = ep.insumo_id
        WHERE g.proyecto_id = ? AND g.activo = 1
    """, (proyecto_id,))

    generadores = sorted(cur.fetchall(), key=lambda r: (_wbs_sort_key(r["wbs"]), r["nombre"] or "", r["id"] or 0))

    for gen in generadores:
        unidad = gen["unidad"] or ""
        desc = gen["concepto_descripcion"] or gen["nombre"] or "SIN DESCRIPCIÓN"
        if gen["wbs"]:
            desc = f"[{gen['wbs']}] {desc}"
        block_start = row_num
        # ── Fila CONCEPTO ──
        ws.cell(row=row_num, column=1, value="Unidad:").font = _LABEL_FONT
        ws.cell(row=row_num, column=2, value=unidad).font = _TITLE_FONT
        ws.cell(row=row_num, column=3, value=desc).font = _TITLE_FONT
        ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=NC)
        for c in [1, 2]:
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_num, column=3).alignment = Alignment(wrap_text=True, vertical="center")
        for c in range(1, NC + 1):
            ws.cell(row=row_num, column=c).fill = _CONCEPTO_FILL
            ws.cell(row=row_num, column=c).border = _THIN_BORDER
        est = len(str(desc)) // 45 + 1
        ws.row_dimensions[row_num].height = max(20, est * 12.5)
        row_num += 1

        # ── Encabezados (sin columna Unidad) ──
        headers = ["Eje", "Tramo", "Largo", "Ancho",
                    "Alto", "No de Pzas.", "Parcial", "Observaciones"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=ci, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _THIN_BORDER
        row_num += 1

        # ── Renglones ──
        cur2 = conn.cursor()
        cur2.execute("""
            SELECT eje, tramo, veces, largo, ancho, alto, subtotal, notas
            FROM generador_renglones
            WHERE generador_id = ? AND activo = 1
            ORDER BY orden
        """, (gen["id"],))

        total = 0.0
        for r in cur2.fetchall():
            subtotal = r["subtotal"] or 0.0
            total += subtotal

            ws.cell(row=row_num, column=1, value=_neutralise(r["eje"]))
            ws.cell(row=row_num, column=2, value=_neutralise(r["tramo"]))

            ws.cell(row=row_num, column=3, value=r["largo"])
            ws.cell(row=row_num, column=3).number_format = _NUM_FMT_4
            ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="right")

            ws.cell(row=row_num, column=4, value=r["ancho"])
            ws.cell(row=row_num, column=4).number_format = _NUM_FMT_4
            ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="right")

            ws.cell(row=row_num, column=5, value=r["alto"])
            ws.cell(row=row_num, column=5).number_format = _NUM_FMT_4
            ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="right")

            ws.cell(row=row_num, column=6, value=r["veces"])
            ws.cell(row=row_num, column=6).number_format = _NUM_FMT
            ws.cell(row=row_num, column=6).alignment = Alignment(horizontal="right")

            ws.cell(row=row_num, column=7, value=round(subtotal, 4))
            ws.cell(row=row_num, column=7).number_format = _NUM_FMT_4
            ws.cell(row=row_num, column=7).alignment = Alignment(horizontal="right")

            ws.cell(row=row_num, column=8, value=_neutralise(r["notas"]))
            row_num += 1

        # ── Fila TOTAL con color ──
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=1, value="").border = _THIN_BORDER
        for c in range(1, NC + 1):
            ws.cell(row=row_num, column=c).fill = _TOTAL_FILL
            ws.cell(row=row_num, column=c).border = _THIN_BORDER
        ws.cell(row=row_num, column=7, value="TOTAL:").font = _TOTAL_FONT
        ws.cell(row=row_num, column=7).fill = _TOTAL_FILL
        ws.cell(row=row_num, column=7).alignment = Alignment(horizontal="right")
        ws.cell(row=row_num, column=8, value=round(total, 4)).font = _TOTAL_FONT
        ws.cell(row=row_num, column=8).fill = _TOTAL_FILL
        ws.cell(row=row_num, column=8).number_format = _NUM_FMT_4
        block_end = row_num
        for r in range(block_start, block_end + 1):
            for c in range(1, NC + 1):
                sides = []
                if r == block_start:
                    sides.append("top")
                if r == block_end:
                    sides.append("bottom")
                if c == 1:
                    sides.append("left")
                if c == NC:
                    sides.append("right")
                border = Border(
                    **{s: Side(style="medium", color="4472C4") for s in sides}
                )
                ws.cell(row=r, column=c).border = border
        row_num += 3  # 2 filas vacías entre bloques

    _auto_width(ws)


def exportar_generadores_excel(
    conn: sqlite3.Connection,
    proyecto_id: int,
    output_path: str,
) -> str:
    wb = Workbook()

    ws_catalogo = wb.active
    ws_catalogo.title = "Catalogo de conceptos"
    _write_catalogo_sheet(ws_catalogo, conn, proyecto_id)

    ws_generadores = wb.create_sheet("Generadores")
    _write_generadores_sheet(ws_generadores, conn, proyecto_id)

    wb.save(output_path)

    return str(Path(output_path).resolve())
